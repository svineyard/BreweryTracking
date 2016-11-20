#!/usr/bin/python

# Program: brewerylistparse.py
# Purpose: parse through brewery list copied and pasted from website and
#          organize data into format appliable to openpyxl use. Then output data
#          to an excel file  

from sys import argv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter 
from unicodetoascii import unicodetoascii

# Read in input brewery text file and output parsed data to new text file
script, readFile, writeFile = argv

# Open brewery list file in read mode
breweries = open(readFile, 'r').read()
output = open(writeFile, 'r+')

# Create a list of lines from brewery list file
brewList = breweries.splitlines()

#initialize list var for parsing
brewListsplit = []

# parse through brewery list file and do:
#   1) Remove initial spacing
#   2) Split brewery name and location 
#   3) Output data to excel sheet 

for i in range(0,len(brewList)):
    temp = brewList[i].split()
    temp = temp[1:]

    # first join list items back into string, then split at parentheses 
    brewList[i] = (' '.join(temp)).split('(')

    # remove (, )," characters from brewery locations (item 2 of each line)
    # translate unicode text to ascii
    brewList[i][1] = brewList[i][1].replace('(','')
    brewList[i][1] = brewList[i][1].replace(')','')  
    brewList[i][0] = unicodetoascii(brewList[i][0])
    brewList[i][1] = unicodetoascii(brewList[i][1])
    brewList[i][1] = brewList[i][1].replace('"','')

    # write to output file parsed brewery list
    #output.write(str(brewList[i]))
    #output.write('\n')

# Create instance of Workbook and new worksheet 
    wb = Workbook()
    ws = wb.active
    ws.title = "Brewery List"

# write to excel doc brewery list
for i in range(0,len(brewList)):
    for j in range(0,2):
        ws.cell(row=i+1, column=j+1,value = brewList[i][j])
        print ws.cell(row=i+1, column=j+1)

# set max column width to change column size in excel
column_widths = []
for row in brewList:
    for i, cell in enumerate(row):
        if len(column_widths) > i:
            if len(cell) > column_widths[i]:
                    column_widths[i] = len(cell)
        else:
            column_widths += [len(cell)]

for i, column_width in enumerate(column_widths):
    ws.column_dimensions[get_column_letter(i+1)].width = column_width


# print(ws.value)
wb.save('Brewery List.xlsx')
        


output.close()





